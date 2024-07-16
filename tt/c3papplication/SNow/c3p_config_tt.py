import mysql.connector
from mysql.connector import Error
from mysql.connector import errorcode
import json
import os
import openpyxl
import requests
import logging
import requests
import time
from c3papplication.common import Connections
from c3papplication.conf.springConfig import springConfig
from datetime import datetime
from werkzeug.utils import secure_filename
import random
import string

configs = springConfig().fetch_config()
logger = logging.getLogger(__name__)

def deviceinfo(mgmtip):
    mydb = Connections.create_connection()
    try:
        mycursor = mydb.cursor(buffered=True)
        dvc_sql = "SELECT d_id FROM c3p_deviceinfo WHERE d_mgmtip = %s or d_hostname = %s"
        logger.debug("c3p_config_tt:deviceinfo :: dvc_sql: %s", dvc_sql)
        mycursor.execute(dvc_sql, (mgmtip, mgmtip,))
        dvc_detail = mycursor.fetchone()
        logger.debug("c3p_config_tt:deviceinfo :: dvc_detail: %s", dvc_detail)
    except Exception as err:
        logger.error("c3p_config_tt:deviceinfo :: Error: %s", err)
    finally:
        mydb.close()
    return dvc_detail


def charactersticinfo(f_name):
    mydb = Connections.create_connection()
    try:
        mycursor = mydb.cursor(buffered=True)
        f_sql = "SELECT f_id FROM c3p_m_features where f_name = %s"
        mycursor.execute(f_sql, (f_name,))
        f_id = mycursor.fetchone()
        if f_id != None:
            cha_sql = "SELECT c_id,c_name FROM c3p_m_characteristics where c_f_id = %s"
            logger.debug("c3p_config_tt:charactersticinfo :: dvc_sql: %s", cha_sql)
            mycursor.execute(cha_sql, (f_id[0],))
            cha_detail = mycursor.fetchall()
            logger.debug("c3p_config_tt:charactersticinfo :: cha_detail: %s", cha_detail)
        else:
            cha_detail = None
            f_id = None
            logger.debug("Feature Not Available: %s", f_name)
    except Exception as err:
        logger.error("c3p_config_tt:charactersticinfo :: Error: %s", err)
    finally:
        mydb.close()
    return cha_detail, f_id

def create_import_id():
    #return ("IRBP"+ (datetime.now().strftime('%Y%m%d%H%M%S%f')[0:16]))
    return ("IMMIGR" + (datetime.now().strftime('%Y%m%d%H%M%S%f')[0:14]) + ''.join(
            random.choices(string.ascii_uppercase, k=1)))

def config_MPS(excel_file):
    try:
        mydb = Connections.create_connection()
        mycursor = mydb.cursor(buffered=True)

        config_json = {}
        config_respns = {}
        i = 0
        fchar = []
        fc_dict_nt = []
        fc_dict = {}
        fc_dict_pcnt = []
        parent_sheet_name = 'MPS_Details'
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        filename = secure_filename(excel_file.filename)
        importId = create_import_id()
        dtime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        updated_by = ""

        if workbook:
            file_status = "Pass"
        else:
            file_status = "Fail"

        if file_status:
            import_type = "Migration"
            sql = "INSERT INTO c3p_m_ds_import_detail(id_created_by,id_created_on,id_name,id_status,id_import_id,id_type) VALUES (%s,%s,%s,%s,%s,%s)"
            logger.debug("ran:c3p_ran:createImportRecordTransactionTable::sql - %s", sql)
            val = (updated_by, dtime, filename, file_status, importId, import_type)
            mycursor.execute(sql, val)
            mydb.commit()

        sheet = workbook[parent_sheet_name]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            fc_Values = []
            sheet = workbook[parent_sheet_name]
            logger.debug("c3p_config_tt:config_MPS :: sheet: %s", sheet)
            variables = [cell.value for cell in sheet[1]]
            logger.debug("c3p_config_tt:config_MPS :: variables_parent: %s", variables)
            device_info = dict(zip(variables, row))
            logger.debug("c3p_config_tt:config_MPS :: device_info_parent: %s", device_info)
            if device_info != None:
                mps_type = device_info['MPS_Type']  # feature name
                mps_host = device_info['exch_mdf_id']
                logger.debug("c3p_config_tt:config_MPS :: mps_type: %s, mps_host: %s", device_info, mps_host)
                deviceinf = deviceinfo(mps_host)  # d_id
                chainfo = charactersticinfo(mps_type)  # char id , char name
                if chainfo[0] != None and deviceinf != None:
                    fc_Values.append(mps_type)
                    for fcchar in chainfo[0]:
                        value = device_info.get(fcchar[1])
                        if (value == ""):
                            value = "None"
                        elif (value == None):
                            value = "None"
                        fc = {
                            "id": fcchar[0],
                            "name": fcchar[1],
                            "value": value,
                            "valueType": "string"
                        }
                        i += 1
                        fchar.append(fc)

                    config_parent_json = {
                        "resourceRelationship": [
                            {
                                "resource": {
                                    "id": deviceinf[0],
                                    "operationalState": "operational",
                                    "activationFeature": [
                                        {
                                            "id": "Copy:::1:::" + chainfo[1][0],
                                            "name": mps_type,
                                            "isBundle": False,
                                            "featureCharacteristic": fchar
                                        }
                                    ]
                                },
                                "relationshipType": "contains"
                            }
                        ]
                    }

                    config_json = json.dumps(config_parent_json)
                    logger.debug("c3p_config_tt:config_MPS :: config_parent_json: %s", config_parent_json)
                    child_sheet_name = 'ExchangeDevice'
                    fpc_char = []
                    j = 0
                    sheet = workbook[child_sheet_name]
                    variables = [cell.value for cell in sheet[1]]

                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        device_info = dict(zip(variables, row))
                        if device_info["Device_Type"] != None:
                            child_type = device_info['Device_Type']
                            mps_child_feature = mps_type + "_" + child_type
                            cha_pc_info = charactersticinfo(mps_child_feature)
                            logger.debug("c3p_config_tt:config_MPS :: cha_pc_info: %s", cha_pc_info)
                            if cha_pc_info[0] != None:
                                fc_Values.append(mps_child_feature)
                                for fpcchar in cha_pc_info[0]:
                                    value = device_info.get(fpcchar[1])
                                    if (value == ""):
                                        value = "None"
                                    elif (value == None):
                                        value = "None"
                                    fpc = {
                                        "id": fpcchar[0],
                                        "name": fpcchar[1],
                                        "value": value,
                                        "valueType": "string"
                                    }
                                    j += 1
                                    fpc_char.append(fpc)

                                configps_json = {
                                    "id": "Copy:::1:::" + cha_pc_info[1][0],
                                    "name": mps_child_feature,
                                    "isBundle": False,
                                    "featureCharacteristic": fpc_char
                                }
                                parent_json = []
                                parent_json = config_parent_json["resourceRelationship"][0]["resource"][
                                    "activationFeature"]
                                parent_json.append(configps_json)
                                logger.debug("c3p_config_tt:config_MPS :: configps_json: %s", configps_json)
                                config_json = json.dumps(config_parent_json)
                                logger.debug("c3p_config_tt:config_MPS :: parent_config_json: %s", config_json)
                            else:
                                fc_dict_pcnt.append(mps_child_feature)

                    logger.info("Before raising request")
                    newHeaders = {"Content-type": "application/json", "Accept": "application/json"}
                    url = configs.get("Python_Application") + "/c3p-p-core/api/ResourceFunction/v4/"
                    config_respns = requests.patch(url, data=config_json, headers=newHeaders, verify=False)
                    config_respns = config_respns.json()
                    id = (config_respns["href"].split("="))[-1]
                    fc_dict[id] = fc_Values
                    logger.debug("c3p_config_tt:config_MPS :: config_respnse_after_raising Request: %s", config_respns)
                    status_query = "SELECT od_requeststatus FROM c3p_rfo_decomposed where od_rfo_id = %s"
                    logger.debug("c3p_config_tt:config_MPS :: status_sql: %s", status_query)
                    mycursor.execute(status_query, (id,))
                    test_status = mycursor.fetchone()
                    count = 0

                    while count <= 36:
                        mydb = Connections.create_connection()
                        mycursor = mydb.cursor(buffered=True)
                        if test_status[0] == "Success" or test_status[0] == "Failure":
                            if test_status[0] == "Success":
                                child_response = config_child(excel_file)
                                response = {"Feature": fc_dict,
                                            "ParentFeature Not Available": fc_dict_nt,
                                            "ParentChildFeature Not Available": fc_dict_pcnt,
                                            "child_response": child_response}
                                logger.debug('c3p_config_tt:config_MPS inside while success :: Response: %s',
                                             response)
                                break
                            else:
                                response = {"Feature": {test_status[0]: fc_Values},
                                            "ParentFeature Not Available": fc_dict_nt,
                                            "ParentChildFeature Not Available": fc_dict_pcnt}
                                logger.debug('c3p_config_tt:config_MPS :: Response: %s', response)
                                break

                        status_query = "SELECT od_requeststatus FROM c3p_rfo_decomposed where od_rfo_id =%s"
                        logger.debug("c3p_config_tt:config_MPS :: status_sql_while: %s", status_query)
                        mycursor.execute(status_query, (id,))
                        test_status = mycursor.fetchone()
                        logger.debug("c3p_config_tt:config_MPS :: test_status_while: %s", test_status[0])
                        time.sleep(5)
                        mydb.close()
                        count += 1
                else:
                    fc_dict_nt.append(mps_type)

    except Exception as e:
        response = jsonify({"Status":"Failure", "Message":"An Internal Error has Occurred"})
        logger.error("c3p_config_tt:config_MPS :: Error In Config_MPS_Parent: %s", e)
    finally:
        mydb.close()
        workbook.close()
    logger.debug("c3p_config_tt:config_MPS :: Response: %s", response)
    return response


def config_child(excel_file_path):
    try:
        mydb = Connections.create_connection()
        mycursor = mydb.cursor(buffered=True)
        config_respns = {}
        i = 0
        fc_Values = []
        fc_dict = {}
        fchild_dict_nt = []
        child_config_json = {}
        logger.debug("Inside Child Config Method")
        sheet_name = 'ExchangeDevice'
        workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
        sheet = workbook[sheet_name]
        variables = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            fchar = []
            device_info = dict(zip(variables, row))
            mps_type = device_info['Device_Type']
            mps_host = device_info['Device_HostName']
            if mps_type != None:
                logger.debug("c3p_config_tt:config_child :: mps_type: %s, mps_host: %s", mps_type, mps_host)
                deviceinf = deviceinfo(mps_host)
                chainfo = charactersticinfo(mps_type)
                if chainfo[0] != None and deviceinf != None:
                    fc_Values.append(mps_type)
                    for fcchar in chainfo[0]:
                        value = device_info.get(fcchar[1])
                        if value == "":
                            value = "None"
                        elif value == None:
                            value = "None"

                        fc = {
                            "id": fcchar[0],
                            "name": fcchar[1],
                            "value": value,
                            "valueType": "string"
                        }
                        i += 1
                        fchar.append(fc)
                    config_json = {
                        "resourceRelationship": [
                            {
                                "resource": {
                                    "id": deviceinf[0],
                                    "operationalState": "operational",
                                    "activationFeature": [
                                        {
                                            "id": "Copy:::1:::" + chainfo[1][0],
                                            "name": mps_type,
                                            "isBundle": False,
                                            "featureCharacteristic": fchar
                                        }
                                    ]
                                },
                                "relationshipType": "contains"
                            }
                        ]
                    }
                    child_config_json = json.dumps(config_json)
                    logger.debug("c3p_config_tt:config_child :: child_config_json: %s", child_config_json)

                    newHeaders = {"Content-type": "application/json", "Accept": "application/json"}
                    url = configs.get("Python_Application") + "/c3p-p-core/api/ResourceFunction/v4/"
                    config_respns = requests.patch(url, data=child_config_json, headers=newHeaders, verify=False)
                    config_respns = config_respns.json()
                    id = (config_respns["href"].split("="))[-1]
                    fc_dict[mps_type] = id
                    logger.debug("c3p_config_tt:config_MPS :: config_respnse: %s", config_respns)
                else:
                    fchild_dict_nt.append(mps_type)

        for i in fc_dict:
            sql = "SELECT od_requeststatus FROM c3p_rfo_decomposed where od_rfo_id = %s"
            logger.debug("c3p_config_tt:config_child :: sql: %s", sql)
            mycursor.execute(sql, (fc_dict[i],))
            status = mycursor.fetchone()
            fc_dict[i] = status[0]

        logger.debug("c3p_config_tt:config_child :: fc_dict_child: %s", fc_dict)
        response = {"Child_Feature": fc_dict, "ChildFeature Not Available": fchild_dict_nt}
    except Exception as e:
        response = str(e)
        logger.error("c3p_config_tt:config_child :: Error: %s", e)
    finally:
        workbook.close()
    return response
